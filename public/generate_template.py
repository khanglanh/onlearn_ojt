#!/usr/bin/env python3
"""
Generate Excel import template based on ExcelParser.java structure
Sheets: Students, Teachers, Courses, Classes, Enrollments
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_import_template():
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. STUDENTS SHEET (removed classId column)
    # Columns: email, name, phoneNumber, studentCode, major, cohort
    students_sheet = wb.create_sheet("Students")
    students_headers = [
        "email", "name", "phoneNumber", "studentCode", "major", "cohort"
    ]
    students_sheet.append(students_headers)
    
    # Sample data for Students
    students_sheet.append([
        "student1@example.com", "John Doe", "+84901234567", "ST001", "Computer Science", "K2024"
    ])
    students_sheet.append([
        "student2@example.com", "Jane Smith", "+84907654321", "ST002", "Software Engineering", "K2024"
    ])
    students_sheet.append([
        "student3@example.com", "Mike Johnson", "+84903456789", "ST003", "Data Science", "K2024"
    ])
    
    # Style Students header
    for col_num, header in enumerate(students_headers, 1):
        cell = students_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        students_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22
    
    # 2. TEACHERS SHEET
    # Columns: email, name, phoneNumber, teacherCode, specialization, department
    teachers_sheet = wb.create_sheet("Teachers")
    teachers_headers = [
        "email", "name", "phoneNumber", "teacherCode", "specialization", "department"
    ]
    teachers_sheet.append(teachers_headers)
    
    # Sample data for Teachers
    teachers_sheet.append([
        "teacher1@example.com", "Dr. Alice Johnson", "+84909876543", "TC001", "Machine Learning", "Computer Science"
    ])
    teachers_sheet.append([
        "teacher2@example.com", "Prof. Bob Wilson", "+84905432198", "TC002", "Web Development", "Software Engineering"
    ])
    
    # Style Teachers header
    for col_num, header in enumerate(teachers_headers, 1):
        cell = teachers_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        teachers_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22
    
    # 3. COURSES SHEET
    # Columns: courseCode, courseName, description, duration, durationInSessions, level, price
    courses_sheet = wb.create_sheet("Courses")
    courses_headers = [
        "courseCode", "courseName", "description", "duration", "durationInSessions", "level", "price"
    ]
    courses_sheet.append(courses_headers)
    
    # Sample data for Courses
    courses_sheet.append([
        "CS101", "Introduction to Programming", "Basic programming concepts using Python", 12, 24, "BEGINNER", 5000000
    ])
    courses_sheet.append([
        "CS201", "Advanced Data Structures", "In-depth study of data structures and algorithms", 16, 32, "INTERMEDIATE", 7500000
    ])
    
    # Style Courses header
    for col_num, header in enumerate(courses_headers, 1):
        cell = courses_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        courses_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 24
    
    # 4. CLASSES SHEET
    # Columns: courseCode, className, teacherEmail, schedule, room, startTime, durationPerSession, capacity, startDate, endDate, enrollKey
    classes_sheet = wb.create_sheet("Classes")
    classes_headers = [
        "courseCode", "className", "teacherEmail", "schedule", "room", "startTime", 
        "durationPerSession", "capacity", "startDate", "endDate", "enrollKey"
    ]
    classes_sheet.append(classes_headers)
    
    # Sample data for Classes
    classes_sheet.append([
        "CS101", "CS101-01-2024", "teacher1@example.com", "Mon,Wed,Fri", "A101", "09:00", 
        90, 30, "2024-09-01", "2024-12-15", "ENROLL2024CS101"
    ])
    classes_sheet.append([
        "CS201", "CS201-01-2024", "teacher2@example.com", "Tue,Thu", "B201", "14:00", 
        120, 25, "2024-09-01", "2024-12-20", "ENROLL2024CS201"
    ])
    
    # Style Classes header
    for col_num, header in enumerate(classes_headers, 1):
        cell = classes_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        classes_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    # 5. ENROLLMENTS SHEET (NEW)
    # Columns: studentEmail, courseCode_className, status, enrolledAt, completedAt
    enrollments_sheet = wb.create_sheet("Enrollments")
    enrollments_headers = [
        "studentEmail", "courseCode_className", "status", "enrolledAt", "completedAt"
    ]
    enrollments_sheet.append(enrollments_headers)
    
    # Sample data for Enrollments showing various relationships:
    # Status values:
    #   * PENDING - Chờ xác nhận (waiting for student to enroll with enrollKey)
    #   * ENROLLED - Đã enroll (after student confirms with enrollKey)
    # - student1: enrolled in CS101 (ENROLLED)
    # - student2: enrolled in CS101 (ENROLLED) and CS201 (PENDING_ACTIVATION)
    # - student3: enrolled in CS201 (PENDING_ACTIVATION)
    # This demonstrates:
    #   * CS101 class has 2 students (student1, student2)
    #   * CS201 class has 2 students (student2, student3)
    #   * student2 is taking 2 classes
    #   * teacher1 teaches CS101 with 2 students
    #   * teacher2 teaches CS201 with 2 students
    enrollments_sheet.append([
        "student1@example.com", "CS101_CS101-01-2024", "ENROLLED", "2024-09-01T08:00:00Z", ""
    ])
    enrollments_sheet.append([
        "student2@example.com", "CS101_CS101-01-2024", "ENROLLED", "2024-09-01T08:30:00Z", ""
    ])
    enrollments_sheet.append([
        "student2@example.com", "CS201_CS201-01-2024", "PENDING", "2024-09-01T09:00:00Z", ""
    ])
    enrollments_sheet.append([
        "student3@example.com", "CS201_CS201-01-2024", "PENDING", "2024-09-01T09:30:00Z", ""
    ])
    
    # Style Enrollments header
    for col_num, header in enumerate(enrollments_headers, 1):
        cell = enrollments_sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        enrollments_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 26
    
    # Save workbook
    wb.save("import-template.xlsx")
    print("✓ Created import-template.xlsx with 5 sheets:")
    print("  - Students (6 columns, 3 sample rows)")
    print("  - Teachers (6 columns, 2 sample rows)")
    print("  - Courses (7 columns, 2 sample rows)")
    print("  - Classes (11 columns, 2 sample rows)")
    print("  - Enrollments (5 columns, 4 sample rows)")
    print("\n✓ Sample data demonstrates complete relationships:")
    print("  - CS101 class: 2 students (student1, student2)")
    print("  - CS201 class: 2 students (student2, student3)")
    print("  - student2 enrolled in 2 classes")
    print("  - teacher1 teaches CS101")
    print("  - teacher2 teaches CS201")

if __name__ == "__main__":
    create_import_template()
